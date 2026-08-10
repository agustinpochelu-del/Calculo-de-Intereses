import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

import leer_mail
import veps
from planillas import COLUMNAS_CAPITAL, COLUMNAS_INTERESES, armar_planilla

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(page_title="Liquidador ARCA — Estudio Pochelú", page_icon="⚖️", layout="wide")

# =====================================================================================
# IDENTIDAD VISUAL
#
# Mismos tokens que la landing del estudio (estudiopochelu.com): verde profundo y
# dorado, Inter para el texto y Source Serif 4 para los títulos, botones tipo píldora
# y fondo crema. Si allá cambia un color, acá hay que cambiarlo también.
# =====================================================================================
st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Source+Serif+4:opsz,wght@8..60,400;8..60,600;8..60,700&display=swap" rel="stylesheet">
<style>
:root {
  --green-900: #082622;
  --green-800: #0E3B34;
  --green-700: #145046;
  --green-600: #1B6B5C;
  --green-500: #2C8A76;
  --gold: #C9A227;
  --gold-soft: #D9BC6A;
  --ink: #0C1A17;
  --body: #445350;
  --muted: #6B7B77;
  --line: #E1E4DF;
  --cream: #F6F5F0;
  --white: #FFFFFF;
  --radius: 14px;
  --radius-lg: 22px;
  --shadow-sm: 0 1px 2px rgba(8,38,34,.06), 0 4px 12px rgba(8,38,34,.05);
  --shadow-md: 0 2px 6px rgba(8,38,34,.07), 0 18px 40px rgba(8,38,34,.09);
  --sans: 'Inter', system-ui, -apple-system, 'Segoe UI', Roboto, sans-serif;
  --serif: 'Source Serif 4', Georgia, 'Times New Roman', serif;
}
.stApp { background: var(--cream); }
html, body, [class*="css"], .stApp, [data-testid="stMarkdownContainer"] {
  font-family: var(--sans);
  color: var(--body);
}
[data-testid="stHeader"] { background: transparent; }
.block-container { padding-top: 2.2rem; max-width: 1240px; }
h1, h2, h3, h4 {
  font-family: var(--serif) !important;
  color: var(--ink) !important;
  letter-spacing: -.015em;
  font-weight: 600 !important;
}
[data-testid="stMarkdownContainer"] strong { color: var(--ink); font-weight: 600; }
/* ── Encabezado de marca ───────────────────── */
.lq-header {
  display: flex; align-items: center; gap: 1rem; flex-wrap: wrap;
  padding: 1.5rem 1.9rem;
  border-radius: var(--radius-lg);
  background:
    radial-gradient(760px 340px at 88% 6%, rgba(44,138,118,.24), transparent 62%),
    linear-gradient(165deg, var(--green-900) 0%, var(--green-800) 58%, #0A322C 100%);
  box-shadow: var(--shadow-md);
  margin-bottom: 1.7rem;
}
.lq-mark {
  display: grid; place-items: center; flex: 0 0 auto;
  width: 46px; height: 46px; border-radius: 13px;
  background: rgba(217,188,106,.16); color: var(--gold-soft); font-size: 1.35rem;
}
.lq-titles { display: flex; flex-direction: column; line-height: 1.2; }
.lq-titles strong {
  font-family: var(--serif); font-size: 1.45rem; font-weight: 700;
  color: #fff; letter-spacing: -.015em;
}
.lq-titles small {
  font-size: .72rem; letter-spacing: .16em; text-transform: uppercase;
  color: rgba(255,255,255,.58); margin-top: .18rem;
}
.lq-eyebrow {
  margin-left: auto; display: inline-flex; align-items: center; gap: .55rem;
  font-size: .74rem; font-weight: 600; letter-spacing: .13em; text-transform: uppercase;
  color: var(--gold-soft);
}
.lq-eyebrow .dot { width: 7px; height: 7px; border-radius: 50%; background: var(--gold); }
/* ── Lengüetas ─────────────────────────────── */
.stTabs [data-baseweb="tab-list"] { gap: .4rem; border-bottom: 1px solid var(--line); }
.stTabs [data-baseweb="tab"] {
  font-family: var(--sans); font-size: .96rem; font-weight: 600;
  color: var(--muted); padding: .55rem 1.15rem; border-radius: 12px 12px 0 0;
}
.stTabs [data-baseweb="tab"]:hover { color: var(--green-700); background: rgba(27,107,92,.06); }
.stTabs [aria-selected="true"] { color: var(--green-800) !important; }
.stTabs [data-baseweb="tab-highlight"] { background: var(--green-600); height: 3px; }
/* ── Métricas ──────────────────────────────── */
[data-testid="stMetric"] {
  background: var(--white); border: 1px solid var(--line); border-left: 4px solid var(--gold);
  border-radius: var(--radius); padding: 1.05rem 1.25rem; box-shadow: var(--shadow-sm);
  transition: transform .25s ease, box-shadow .25s ease;
}
[data-testid="stMetric"]:hover { transform: translateY(-3px); box-shadow: var(--shadow-md); }
[data-testid="stMetricLabel"] p {
  font-size: .74rem !important; font-weight: 600; letter-spacing: .1em;
  text-transform: uppercase; color: var(--muted) !important;
}
[data-testid="stMetricValue"] {
  font-family: var(--serif); font-size: 1.6rem !important;
  color: var(--green-800) !important; font-weight: 600;
}
/* ── Botones ───────────────────────────────── */
.stButton button, .stDownloadButton button, [data-testid="stBaseButton-secondary"] {
  font-family: var(--sans); font-weight: 600; font-size: .95rem;
  border-radius: 999px !important; padding: .75rem 1.5rem;
  border: 1.5px solid transparent !important;
  transition: transform .18s ease, background-color .18s ease, box-shadow .18s ease;
}
.stButton button:hover, .stDownloadButton button:hover { transform: translateY(-2px); }
.stDownloadButton button {
  background: var(--green-800) !important; color: #fff !important;
  box-shadow: var(--shadow-sm);
}
.stDownloadButton button:hover { background: var(--green-700) !important; box-shadow: var(--shadow-md); }
.stDownloadButton button p { color: #fff !important; font-weight: 600; }
/* La plantilla en blanco es la acción secundaria: dorada, no verde. */
.st-key-lq-plantilla-capital .stDownloadButton button,
.st-key-lq-plantilla-intereses .stDownloadButton button {
  background: var(--gold-soft) !important; box-shadow: none;
}
.st-key-lq-plantilla-capital .stDownloadButton button:hover,
.st-key-lq-plantilla-intereses .stDownloadButton button:hover { background: #E6CE86 !important; }
.st-key-lq-plantilla-capital .stDownloadButton button p,
.st-key-lq-plantilla-intereses .stDownloadButton button p { color: var(--green-900) !important; }
/* ── Carga de archivo ──────────────────────── */
[data-testid="stFileUploaderDropzone"] {
  background: var(--white); border: 1.5px dashed #C6D1CD; border-radius: var(--radius);
  padding: 1.4rem;
}
[data-testid="stFileUploaderDropzone"]:hover { border-color: var(--green-500); background: rgba(27,107,92,.03); }
[data-testid="stFileUploaderDropzone"] button {
  border-radius: 999px !important; border: 1.5px solid rgba(12,26,23,.22) !important;
  background: transparent !important; color: var(--ink) !important; font-weight: 600;
}
[data-testid="stFileUploaderDropzone"] button:hover {
  border-color: var(--green-700) !important; color: var(--green-800) !important;
  background: rgba(27,107,92,.06) !important;
}
/* ── Avisos ────────────────────────────────── */
/* El fondo lo pinta stAlertContainer, no stAlert: el tipo de aviso se detecta con
   :has() sobre el testid del contenido, que es donde Streamlit lo marca. */
[data-testid="stAlert"] { border: 0; box-shadow: none; padding: 0; }
[data-testid="stAlertContainer"] {
  border-radius: var(--radius) !important; border: 1px solid var(--line);
  border-left: 4px solid var(--green-500); box-shadow: var(--shadow-sm);
  color: var(--body);
}
[data-testid="stAlertContainer"]:has([data-testid="stAlertContentSuccess"]) {
  background: rgba(44,138,118,.10) !important; border-left-color: var(--green-600);
}
[data-testid="stAlertContainer"]:has([data-testid="stAlertContentInfo"]) {
  background: rgba(20,80,70,.07) !important; border-left-color: var(--green-500);
}
[data-testid="stAlertContainer"]:has([data-testid="stAlertContentWarning"]) {
  background: rgba(201,162,39,.12) !important; border-left-color: var(--gold);
}
[data-testid="stAlertContainer"]:has([data-testid="stAlertContentError"]) {
  background: rgba(180,35,24,.08) !important; border-left-color: #B42318;
}
/* ── Tablas y desplegables ─────────────────── */
[data-testid="stExpander"] {
  background: var(--white); border: 1px solid var(--line);
  border-radius: var(--radius-lg); box-shadow: var(--shadow-sm); overflow: hidden;
}
[data-testid="stExpander"] summary { font-weight: 600; color: var(--ink); }
[data-testid="stExpander"] summary:hover { color: var(--green-700); }
[data-testid="stDataFrame"] {
  border: 1px solid var(--line); border-radius: var(--radius);
  overflow: hidden; box-shadow: var(--shadow-sm); background: var(--white);
}
hr, [data-testid="stDivider"] hr { border-color: var(--line); }
[data-testid="stCaptionContainer"] p { color: var(--muted); font-size: .87rem; }
:focus-visible { outline: 3px solid var(--green-500); outline-offset: 3px; border-radius: 6px; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="lq-header">
  <span class="lq-mark">⚖</span>
  <span class="lq-titles">
    <strong>Liquidador ARCA</strong>
    <small>Ejecución fiscal</small>
  </span>
  <span class="lq-eyebrow"><span class="dot"></span>Estudio Pochelú</span>
</div>
""", unsafe_allow_html=True)

def formato_arg(numero):
    return f"${numero:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# =====================================================================================
# TASAS: viven en tasas.json, al lado de este archivo
#
# Ese archivo es el ÚNICO lugar donde se actualizan las tasas, y replica la tabla
# "Evolución de Tasas de Intereses" que publica ARCA (una fila por tramo, con la tasa
# resarcitoria y la punitoria juntas, igual que en el original). No hace falta cargar
# tasas en el Excel que se sube: la app ya las tiene, desde 1901 hasta el tramo vigente.
#
# Las tasas se guardan como TASA MENSUAL en porcentaje, que es como las publica ARCA.
# La tasa diaria se deriva dividiendo por 30 en el momento del cálculo, sin truncar
# decimales (0,0275/30 y no 0,00091667: truncar movía centavos en cada línea).
#
# "hasta" es INCLUSIVO: el tramo rige hasta el final de ese día y el siguiente arranca
# al día siguiente, así que hasta + 1 día tiene que dar el "desde" del tramo que sigue.
# Si queda un hueco se pierden días de interés sin que nadie lo note, y por eso la
# continuidad se verifica al cargar.
#
# "dias" son los días que ARCA computa para el tramo COMPLETO, tomados del detalle de
# cálculo que emite. No siempre coinciden con lo que daría la cuenta por meses de 30
# (ver dias_arca): en el bimestre 12/2024-01/2025 ARCA computa 61 y no 60, en febrero
# 2025 computa 28 y no 30, y en 03/2025-06/2025 computa 122 y no 120. Son los días
# oficiales y mandan sobre cualquier fórmula, pero solo cuando el devengamiento cubre
# el tramo entero; en los tramos de las puntas se cuenta con dias_arca. ARCA no los
# publica en la tabla de tasas, así que los tramos sin el dato quedan en null y la app
# avisa en pantalla si le tocó usar uno completo sin él.
#
# 🔧 Para actualizar: correr `python actualizar_tasas.py`, que compara contra la página
# de ARCA. El control automático de .github/workflows/control-tasas.yml lo corre solo
# una vez por semana y avisa si aparece un tramo nuevo.
# =====================================================================================
ARCHIVO_TASAS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tasas.json")


# ARCA cierra el tramo vigente en 2999-12-31, que pandas no puede representar (su
# máximo es abril de 2262). Se recorta ahí: a los efectos del cálculo, cualquier fecha
# tan lejana significa lo mismo, "vigente hasta nuevo aviso".
_TOPE_FECHA = pd.Timestamp('2262-01-01')


def _fecha(iso):
    return _TOPE_FECHA if int(iso[:4]) >= 2262 else pd.Timestamp(iso)


def _tramos_a_df(tramos, campo_tasa):
    """Arma la tabla de un tipo de interés (resarcitorio o punitorio) a partir de los
    tramos de tasas.json, que traen las dos tasas en la misma fila."""
    df = pd.DataFrame([{
        'Desde': _fecha(t['desde']),
        'Hasta': _fecha(t['hasta']),
        'Norma': t.get('norma', ''),
        'Tasa_Mensual': t[campo_tasa],
        'Dias': t.get('dias'),
    } for t in tramos])
    df['Tasa_Mensual'] = pd.to_numeric(df['Tasa_Mensual'])
    df['Tasa_Diaria'] = df['Tasa_Mensual'] / 100 / 30
    df['Dias'] = pd.to_numeric(df['Dias'], errors='coerce')
    return df.sort_values('Desde').reset_index(drop=True)


@st.cache_data
def cargar_tasas():
    """Lee tasas.json y devuelve (resarcitorias, punitorias, metadatos).

    Si el archivo no está o está mal, no hay cálculo posible: se corta con un error
    claro en lugar de liquidar con tasas a medias.
    """
    with open(ARCHIVO_TASAS, encoding='utf-8') as fh:
        doc = json.load(fh)

    tramos = doc['tramos']
    if not tramos:
        raise ValueError("tasas.json no tiene ningún tramo cargado")

    faltan = [c for c in ('desde', 'hasta', 'resarcitoria_mensual', 'punitoria_mensual')
              if any(c not in t for t in tramos)]
    if faltan:
        raise ValueError(f"hay tramos en tasas.json sin los campos: {', '.join(faltan)}")

    df_res = _tramos_a_df(tramos, 'resarcitoria_mensual')
    df_pun = _tramos_a_df(tramos, 'punitoria_mensual')

    huecos = []
    for i in range(len(df_res) - 1):
        esperado = df_res.iloc[i]['Hasta'] + pd.Timedelta(days=1)
        if df_res.iloc[i + 1]['Desde'] != esperado:
            huecos.append(f"{df_res.iloc[i]['Hasta']:%d/%m/%Y} → {df_res.iloc[i + 1]['Desde']:%d/%m/%Y}")
    if huecos:
        raise ValueError("los tramos de tasas.json no empalman: " + "; ".join(huecos))

    meta = {'verificado': doc.get('verificado', 'sin fecha'), 'fuente': doc.get('fuente', ''),
            'tramos': len(tramos)}
    return df_res, df_pun, meta


# =====================================================================================
# MOTOR DE CÁLCULO (compartido por las dos lengüetas)
# =====================================================================================
def dias_arca(desde, hasta):
    """Días entre dos fechas con la convención de ARCA: cada mes completo cuenta 30
    días, y el resto se cuenta por días corridos.

    Es la diferencia clave contra un conteo de días calendario. Ejemplo real, un
    anticipo con vencimiento 13/02/2026 liquidado al 10/06/2026:
        días calendario = 117   (15 de febrero + 31 + 30 + 31 + 10)
        días ARCA       = 118   (3 meses x 30 + 28 días de 13/05 a 10/06)
    Febrero corto suma 30 igual, y por eso el cálculo por días corridos da de menos
    en unos meses y de más en otros.
    """
    if pd.isna(desde) or pd.isna(hasta) or hasta <= desde:
        return 0
    meses = (hasta.year - desde.year) * 12 + (hasta.month - desde.month)
    ancla = desde + pd.DateOffset(months=meses)
    if ancla > hasta:
        meses -= 1
        ancla = desde + pd.DateOffset(months=meses)
    return meses * 30 + (hasta - ancla).days


def calcular_interes(fecha_inicio_calculo, fecha_fin_calculo, capital, df_tabla_tasas, avisos=None):
    """Devuelve (interés, días) del período, aplicando cada tramo de tasa vigente.

    El interés empieza a devengar el día SIGUIENTE a la fecha de origen (el
    vencimiento, la demanda o el pago, según el tramo) y corre hasta la fecha de
    corte inclusive: es lo que muestra el detalle de ARCA, donde un vencimiento del
    13/05/2024 abre el primer tramo el 14/05/2024.

    Para contar los días de cada tramo hay dos casos:

    - El devengamiento cubre el tramo ENTERO: se usan los días oficiales de la
      tabla ("Dias"), que es lo que computa ARCA aunque no siempre coincida con la
      cuenta por meses de 30.
    - El devengamiento cubre el tramo a medias (los tramos de las puntas): se
      cuentan con dias_arca sobre la parte efectivamente cubierta.

    El interés de cada tramo se redondea antes de sumarlo, igual que en el detalle
    de ARCA, donde los importes por tramo suman exacto el total informado.

    Si se pasa un set en `avisos`, se le agregan los tramos que hubo que usar
    completos sin tener cargados los días oficiales, para poder avisarlo en pantalla.
    """
    if pd.isna(fecha_inicio_calculo) or pd.isna(fecha_fin_calculo):
        return 0.0, 0
    if fecha_inicio_calculo >= fecha_fin_calculo or capital == 0:
        return 0.0, 0

    un_dia = pd.Timedelta(days=1)
    # Ventana de devengamiento, con el extremo superior abierto.
    devenga_desde = fecha_inicio_calculo + un_dia
    devenga_hasta = fecha_fin_calculo + un_dia

    interes_acumulado = 0.0
    dias_acumulados = 0

    for _, tramo in df_tabla_tasas.iterrows():
        tramo_desde = tramo['Desde']
        tramo_hasta = tramo['Hasta'] + un_dia
        inicio = max(devenga_desde, tramo_desde)
        fin = min(devenga_hasta, tramo_hasta)
        if inicio >= fin:
            continue

        cubre_tramo_entero = (inicio == tramo_desde) and (fin == tramo_hasta)
        if cubre_tramo_entero and pd.notna(tramo['Dias']):
            dias = int(tramo['Dias'])
        else:
            dias = dias_arca(inicio, fin)
            if cubre_tramo_entero and avisos is not None:
                avisos.add(f"{tramo_desde:%d/%m/%Y} al {tramo['Hasta']:%d/%m/%Y}")
        if dias <= 0:
            continue

        interes_acumulado += round(capital * tramo['Tasa_Diaria'] * dias, 2)
        dias_acumulados += dias

    return round(interes_acumulado, 2), dias_acumulados


def avisar_hojas_de_tasas_ignoradas(archivo_subido):
    """Los archivos armados con las plantillas viejas traen hojas de tasas. Ya no se
    usan: la app tiene la tabla oficial completa, y una hoja desactualizada metida en
    el medio cambiaría un importe que va a un expediente sin que nadie lo note."""
    try:
        hojas = pd.ExcelFile(archivo_subido).sheet_names
    except Exception:
        return
    viejas = [h for h in hojas if h.strip().lower().startswith('tasas')]
    if viejas:
        st.info(
            f"ℹ️ Tu archivo trae la hoja **{'** y **'.join(viejas)}**. No se usa: la app "
            f"liquida con la tabla oficial de ARCA que tiene cargada. Podés borrar esas hojas.")


def avisar_tramos_sin_dias(avisos):
    if avisos:
        st.warning(
            "⚠️ Estos tramos se usaron completos sin tener cargados los días oficiales de "
            "ARCA, así que la app los contó sola (meses de 30 días) y puede quedar un día "
            f"de diferencia: **{'**, **'.join(sorted(avisos))}**. Para dejarlo exacto, pedile "
            "a ARCA un detalle de cálculo que atraviese esos tramos y pasame el número.")


# =====================================================================================
# VALIDACIONES DE LA HOJA "Deudas"
# =====================================================================================
def validar_deudas(df, columnas_requeridas, df_tasas_res, df_tasas_pun):
    """Devuelve una lista de problemas encontrados. Lista vacía = todo en orden."""
    problemas = []

    faltantes = [c for c in columnas_requeridas if c not in df.columns]
    if faltantes:
        problemas.append(f"Faltan columnas en la hoja 'Deudas': {', '.join(faltantes)}.")
        return problemas

    if df.empty:
        problemas.append("La hoja 'Deudas' no tiene ninguna obligación cargada.")
        return problemas

    sin_capital = df['Capital'].isna() | (df['Capital'] <= 0)
    if sin_capital.any():
        filas = ', '.join(str(i + 2) for i in df.index[sin_capital])
        problemas.append(f"Hay capital vacío o menor o igual a cero (filas del Excel: {filas}).")

    mal_ordenadas = df['Vencimiento'] > df['fecha_Demanda']
    if mal_ordenadas.any():
        filas = ', '.join(str(i + 2) for i in df.index[mal_ordenadas])
        problemas.append(f"El vencimiento es posterior a la fecha de demanda (filas del Excel: {filas}).")

    liq_previa = df['Fecha_Liquidacion'] < df['fecha_Demanda']
    if liq_previa.any():
        filas = ', '.join(str(i + 2) for i in df.index[liq_previa])
        problemas.append(f"La fecha de liquidación es anterior a la de demanda (filas del Excel: {filas}).")

    # La tabla de tasas tiene que cubrir todo el período que se liquida. Hoy va de 1901
    # al tramo vigente, así que esto no debería saltar nunca; está por si alguien recorta
    # tasas.json y deja un caso viejo sin tasa, que se liquidaría en cero sin avisar.
    desde_tabla, hasta_tabla = df_tasas_res['Desde'].min(), df_tasas_res['Hasta'].max()
    if df['Vencimiento'].min() < desde_tabla:
        problemas.append(
            f"La tabla de tasas arranca el {desde_tabla:%d/%m/%Y} y hay obligaciones que "
            f"vencen antes ({df['Vencimiento'].min():%d/%m/%Y}): faltan tramos.")
    fin_calculo = max(df['fecha_Demanda'].max(), df['Fecha_Liquidacion'].max())
    if fin_calculo > hasta_tabla:
        problemas.append(
            f"La tabla de tasas termina el {hasta_tabla:%d/%m/%Y} y el cálculo llega hasta "
            f"el {fin_calculo:%d/%m/%Y}: faltan tramos.")

    return problemas


# =====================================================================================
# UI: TABLA DE TASAS DE REFERENCIA
# =====================================================================================
def _recortar(df_tasas, desde, hasta):
    """Deja solo los tramos que el cálculo efectivamente toca, con las puntas recortadas
    a las fechas del juicio. En una punta recortada los días oficiales del tramo completo
    ya no aplican, así que se recalculan sobre el pedazo que quedó."""
    un_dia = pd.Timedelta(days=1)
    df = df_tasas[(df_tasas['Hasta'] >= desde) & (df_tasas['Desde'] <= hasta)].copy()
    if df.empty:
        return df
    ini = df.columns.get_loc('Desde')
    fin = df.columns.get_loc('Hasta')
    dias = df.columns.get_loc('Dias')
    if df.iloc[0]['Desde'] < desde:
        df.iloc[0, ini] = desde
        df.iloc[0, dias] = None
    if df.iloc[-1]['Hasta'] > hasta:
        df.iloc[-1, fin] = hasta
        df.iloc[-1, dias] = None
    df['Días'] = [int(d) if pd.notna(d) else dias_arca(des, has + un_dia)
                  for d, des, has in zip(df['Dias'], df['Desde'], df['Hasta'])]
    return df


def mostrar_tabla_tasas(df_tasas_res, df_tasas_pun, meta, ventana_res, ventana_pun,
                        titulo_res="**Resarcitorios y Capitalizables**"):
    def formatear(df):
        if df.empty:
            return df
        df = df.copy()
        df['Tasa mensual'] = df['Tasa_Mensual'].apply(lambda x: f"{x:g}%".replace('.', ','))
        df['Tasa diaria'] = df['Tasa_Diaria'].apply(lambda x: f"{x*100:.6f}%".replace('.', ','))
        df['Desde'] = df['Desde'].dt.strftime('%d/%m/%Y')
        df['Hasta'] = df['Hasta'].dt.strftime('%d/%m/%Y')
        cols = ['Desde', 'Hasta', 'Tasa mensual', 'Tasa diaria', 'Días']
        if 'Norma' in df.columns:
            cols.append('Norma')
        return df[cols]

    st.markdown("### 📈 Tasas de Interés aplicadas")
    st.caption(
        f"Solo los tramos que toca este cálculo. Tabla oficial de ARCA: {meta['tramos']} tramos "
        f"cargados, verificada el {meta['verificado']}. Los tramos completos usan los días "
        f"oficiales de ARCA; en las puntas los días se cuentan como los cuenta ARCA (cada mes "
        f"completo vale 30 días y el resto por días corridos).")
    col_t1, col_t2 = st.columns(2)

    with col_t1:
        st.write(titulo_res)
        st.dataframe(formatear(_recortar(df_tasas_res, *ventana_res)), use_container_width=True, hide_index=True)

    with col_t2:
        st.write("**Punitorios (desde el inicio de la ejecución)**")
        st.dataframe(formatear(_recortar(df_tasas_pun, *ventana_pun)), use_container_width=True, hide_index=True)


# =====================================================================================
# DESCARGA DEL EXCEL LIQUIDADO
# =====================================================================================
def boton_descarga(df_deudas, nombre_archivo, columnas_moneda, columnas_totalizar,
                   clave=""):
    # `clave` distingue las instancias: la misma liquidacion puede estar en pantalla
    # dos veces (subida en su lengueta y liquidada desde el mail), y Streamlit pide
    # que cada control tenga un nombre propio.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Liquidacion_Apremio'

    columnas = list(df_deudas.columns)
    _estilo_header(ws, 1, columnas)

    for fila_idx, (_, fila) in enumerate(df_deudas.iterrows(), start=2):
        for col_idx, col in enumerate(columnas, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=fila[col])
            celda.font = Font(name=_FUENTE, size=10)
            if col in columnas_moneda:
                celda.number_format = '#,##0.00'

    # Fila de totales, para no tener que sumar a mano al presentar la liquidación.
    # Se escriben importes, no fórmulas: la planilla se presenta en el expediente y
    # tiene que mostrar el total aunque se abra con un lector que no evalúe fórmulas.
    fila_total = len(df_deudas) + 2
    ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(name=_FUENTE, bold=True, size=10)
    for col_idx, col in enumerate(columnas, start=1):
        if col in columnas_totalizar:
            celda = ws.cell(row=fila_total, column=col_idx, value=round(df_deudas[col].sum(), 2))
            celda.font = Font(name=_FUENTE, bold=True, size=10)
            celda.number_format = '#,##0.00'
    for col_idx in range(1, len(columnas) + 1):
        ws.cell(row=fila_total, column=col_idx).fill = PatternFill(
            start_color=_GRIS_EJEMPLO, end_color=_GRIS_EJEMPLO, fill_type="solid")

    _ajustar_anchos(ws, [max(12, min(26, len(str(c)) + 4)) for c in columnas])
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)

    st.markdown("<br>", unsafe_allow_html=True)
    c1, c_boton, c2 = st.columns([1, 2, 1])
    with c_boton:
        st.download_button(
            label="📥 Descargar Planilla (Excel)",
            data=output.getvalue(),
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"{nombre_archivo}_{clave}"
        )


# =====================================================================================
# LENGÜETA 1: JUICIO POR CAPITAL + INTERESES (capital impositivo impago, con o sin pago posterior)
# =====================================================================================
def procesar_juicio_capital(origen, clave="capital"):
    """`origen` es el Excel que se sube, o la planilla ya armada en memoria.

    Lo segundo es lo que usa la importación desde el mail: los datos ya están
    revisados en pantalla, no tiene sentido obligar a bajar el archivo y volver
    a subirlo. El cálculo es el mismo por los dos caminos.
    """
    if isinstance(origen, pd.DataFrame):
        df_deudas = origen.copy()
    else:
        df_deudas = pd.read_excel(origen, sheet_name='Deudas')
        avisar_hojas_de_tasas_ignoradas(origen)
    df_tasas_res, df_tasas_pun, meta = cargar_tasas()

    df_deudas.columns = df_deudas.columns.str.strip()
    df_deudas = df_deudas.dropna(subset=['Vencimiento'])

    if 'F. Pago Capital' not in df_deudas.columns:
        df_deudas['F. Pago Capital'] = pd.NaT

    df_deudas['Vencimiento'] = pd.to_datetime(df_deudas['Vencimiento'])
    df_deudas['F. Pago Capital'] = pd.to_datetime(df_deudas['F. Pago Capital'])
    df_deudas['fecha_Demanda'] = pd.to_datetime(df_deudas['fecha_Demanda'])
    df_deudas['Fecha_Liquidacion'] = pd.to_datetime(df_deudas['Fecha_Liquidacion'])

    problemas = validar_deudas(
        df_deudas, ['Impuesto', 'Vencimiento', 'Capital', 'fecha_Demanda', 'Fecha_Liquidacion'],
        df_tasas_res, df_tasas_pun)
    if problemas:
        for p in problemas:
            st.error(f"❌ {p}")
        return

    ultima_fecha_demanda = df_deudas['fecha_Demanda'].max()
    ultima_fecha_liq = df_deudas['Fecha_Liquidacion'].max()

    # Los tres tramos son continuos, sin salto de un día entre uno y el siguiente.
    # El orden de Punitorios y Capitalizables se invierte según cuál de los dos
    # eventos (Demanda o Pago del Capital) ocurre primero:
    #
    #   Caso A: Demanda ANTES del Pago del Capital
    #       Resarcitorios:  Vencimiento -> Demanda            (sobre Capital)
    #       Punitorios:     Demanda -> Pago Capital            (sobre Capital)
    #       Capitalizables: Pago Capital -> Liquidación         (sobre monto Resarcitorios)
    #
    #   Caso B: Pago del Capital ANTES de la Demanda
    #       Resarcitorios:  Vencimiento -> Pago Capital        (sobre Capital)
    #       Capitalizables: Pago Capital -> Demanda            (sobre monto Resarcitorios)
    #       Punitorios:     Demanda -> Liquidación              (sobre Capital)
    #
    #   Sin fecha de Pago de Capital cargada: Resarcitorios hasta la Demanda,
    #   Punitorios desde la Demanda hasta la Liquidación, sin Capitalizables.
    avisos = set()

    def procesar_fila(fila):
        vencimiento = fila['Vencimiento']
        pago_capital = fila['F. Pago Capital']
        fecha_demanda = fila['fecha_Demanda']
        fecha_liq = fila['Fecha_Liquidacion']
        capital = fila['Capital']

        if pd.isna(pago_capital):
            resarcitorio, dias_res = calcular_interes(vencimiento, fecha_demanda, capital, df_tasas_res, avisos)
            capitalizable, dias_cap = 0.0, 0
            punitorio, dias_pun = calcular_interes(fecha_demanda, fecha_liq, capital, df_tasas_pun, avisos)
        elif fecha_demanda <= pago_capital:
            resarcitorio, dias_res = calcular_interes(vencimiento, fecha_demanda, capital, df_tasas_res, avisos)
            punitorio, dias_pun = calcular_interes(fecha_demanda, pago_capital, capital, df_tasas_pun, avisos)
            capitalizable, dias_cap = calcular_interes(pago_capital, fecha_liq, resarcitorio, df_tasas_res, avisos)
        else:
            resarcitorio, dias_res = calcular_interes(vencimiento, pago_capital, capital, df_tasas_res, avisos)
            capitalizable, dias_cap = calcular_interes(pago_capital, fecha_demanda, resarcitorio, df_tasas_res, avisos)
            punitorio, dias_pun = calcular_interes(fecha_demanda, fecha_liq, capital, df_tasas_pun, avisos)

        return pd.Series({
            'Interes_Resarcitorio': resarcitorio, 'Dias_Resarcitorios': dias_res,
            'Interes_Capitalizable': capitalizable, 'Dias_Capitalizables': dias_cap,
            'Interes_Punitorio': punitorio, 'Dias_Punitorios': dias_pun,
        })

    df_deudas = df_deudas.join(df_deudas.apply(procesar_fila, axis=1))
    for col in ['Dias_Resarcitorios', 'Dias_Capitalizables', 'Dias_Punitorios']:
        df_deudas[col] = df_deudas[col].astype(int)

    fecha_inicio_punitorios_global = ultima_fecha_demanda
    antiguedad_juicio = max(0, (ultima_fecha_liq - fecha_inicio_punitorios_global).days)

    df_deudas['Total_Actualizado'] = (
        df_deudas['Capital'] + df_deudas['Interes_Resarcitorio']
        + df_deudas['Interes_Capitalizable'] + df_deudas['Interes_Punitorio']
    )

    df_deudas_fmt = df_deudas.copy()
    df_deudas_fmt['Vencimiento'] = df_deudas_fmt['Vencimiento'].dt.strftime('%d/%m/%Y')
    df_deudas_fmt['F. Pago Capital'] = df_deudas_fmt['F. Pago Capital'].apply(lambda d: d.strftime('%d/%m/%Y') if pd.notna(d) else '')
    df_deudas_fmt['fecha_Demanda'] = df_deudas_fmt['fecha_Demanda'].dt.strftime('%d/%m/%Y')
    df_deudas_fmt['Fecha_Liquidacion'] = df_deudas_fmt['Fecha_Liquidacion'].dt.strftime('%d/%m/%Y')

    st.success("¡Liquidación judicial finalizada!")
    avisar_tramos_sin_dias(avisos)
    st.markdown("### 📋 Resumen del Juicio")

    col1, col2, col3, col4 = st.columns(4)
    with col1: st.metric("Capital Original", formato_arg(df_deudas['Capital'].sum()))
    with col2: st.metric("Resarcitorios", formato_arg(df_deudas['Interes_Resarcitorio'].sum()))
    with col3: st.metric("Capitalizables", formato_arg(df_deudas['Interes_Capitalizable'].sum()))
    with col4: st.metric("Punitorios", formato_arg(df_deudas['Interes_Punitorio'].sum()))

    col5, col6 = st.columns(2)
    with col5: st.metric("Total Actualizado", formato_arg(df_deudas['Total_Actualizado'].sum()))
    with col6: st.metric("Antigüedad Juicio", f"{antiguedad_juicio} días")

    st.divider()

    columnas_detalle = ['Impuesto', 'Vencimiento', 'Capital', 'F. Pago Capital',
                        'Dias_Resarcitorios', 'Interes_Resarcitorio',
                        'Dias_Capitalizables', 'Interes_Capitalizable',
                        'fecha_Demanda', 'Dias_Punitorios', 'Interes_Punitorio',
                        'Fecha_Liquidacion', 'Total_Actualizado']

    with st.expander("🔍 Ver detalle completo de obligaciones", expanded=False):
        st.dataframe(df_deudas_fmt[columnas_detalle], use_container_width=True)

    st.divider()
    mostrar_tabla_tasas(
        df_tasas_res, df_tasas_pun, meta,
        ventana_res=(df_deudas['Vencimiento'].min(), ultima_fecha_demanda),
        ventana_pun=(df_deudas['fecha_Demanda'].min(), ultima_fecha_liq))

    boton_descarga(
        df_deudas_fmt[columnas_detalle], "Liquidacion_ARCA_Apremio.xlsx", clave=clave,
        columnas_moneda=['Capital', 'Interes_Resarcitorio', 'Interes_Capitalizable',
                         'Interes_Punitorio', 'Total_Actualizado'],
        columnas_totalizar=['Capital', 'Interes_Resarcitorio', 'Interes_Capitalizable',
                            'Interes_Punitorio', 'Total_Actualizado'])

    # Los importes ya están calculados acá: no hace falta bajar la planilla y volver
    # a subirla para pagar. Va plegado porque no siempre se paga en el momento.
    with st.expander("🧾 Generar los VEPs para pagar esto"):
        generar_veps(df_deudas, clave)


# =====================================================================================
# LENGÜETA 2: JUICIO A LOS INTERESES (el capital impositivo ya está pago; se demanda por
# los intereses resarcitorios impagos, que pasan a ser la nueva "base" de la deuda)
# =====================================================================================
def procesar_juicio_intereses(origen, clave="intereses"):
    """Igual que `procesar_juicio_capital`: acepta el Excel o la planilla en memoria."""
    if isinstance(origen, pd.DataFrame):
        df_deudas = origen.copy()
    else:
        df_deudas = pd.read_excel(origen, sheet_name='Deudas')
        avisar_hojas_de_tasas_ignoradas(origen)
    df_tasas_res, df_tasas_pun, meta = cargar_tasas()

    df_deudas.columns = df_deudas.columns.str.strip()
    df_deudas = df_deudas.dropna(subset=['Vencimiento'])

    df_deudas['Vencimiento'] = pd.to_datetime(df_deudas['Vencimiento'])
    df_deudas['fecha_Demanda'] = pd.to_datetime(df_deudas['fecha_Demanda'])
    df_deudas['Fecha_Liquidacion'] = pd.to_datetime(df_deudas['Fecha_Liquidacion'])

    problemas = validar_deudas(
        df_deudas, ['Impuesto', 'Vencimiento', 'Capital', 'fecha_Demanda', 'Fecha_Liquidacion'],
        df_tasas_res, df_tasas_pun)
    if problemas:
        for p in problemas:
            st.error(f"❌ {p}")
        return

    ultima_fecha_demanda = df_deudas['fecha_Demanda'].max()
    ultima_fecha_liq = df_deudas['Fecha_Liquidacion'].max()

    # Acá "Capital" ya es el monto de intereses que se convirtió en la base del juicio
    # (el capital impositivo original está pago, no interviene). Solo hay dos tramos,
    # continuos, sin salto de un día entre uno y el siguiente:
    #   Resarcitorios: Vencimiento -> Demanda   (sobre ese monto base)
    #   Punitorios:    Demanda -> Liquidación   (sobre ese monto base)
    avisos = set()

    def procesar_fila(fila):
        resarcitorio, dias_res = calcular_interes(fila['Vencimiento'], fila['fecha_Demanda'], fila['Capital'], df_tasas_res, avisos)
        punitorio, dias_pun = calcular_interes(fila['fecha_Demanda'], fila['Fecha_Liquidacion'], fila['Capital'], df_tasas_pun, avisos)
        return pd.Series({
            'Interes_Resarcitorio': resarcitorio, 'Dias_Resarcitorios': dias_res,
            'Interes_Punitorio': punitorio, 'Dias_Punitorios': dias_pun,
        })

    df_deudas = df_deudas.join(df_deudas.apply(procesar_fila, axis=1))
    for col in ['Dias_Resarcitorios', 'Dias_Punitorios']:
        df_deudas[col] = df_deudas[col].astype(int)

    fecha_inicio_punitorios_global = ultima_fecha_demanda
    antiguedad_juicio = max(0, (ultima_fecha_liq - fecha_inicio_punitorios_global).days)

    df_deudas['Total_Actualizado'] = df_deudas['Capital'] + df_deudas['Interes_Resarcitorio'] + df_deudas['Interes_Punitorio']

    df_deudas_fmt = df_deudas.copy()
    df_deudas_fmt['Vencimiento'] = df_deudas_fmt['Vencimiento'].dt.strftime('%d/%m/%Y')
    df_deudas_fmt['fecha_Demanda'] = df_deudas_fmt['fecha_Demanda'].dt.strftime('%d/%m/%Y')
    df_deudas_fmt['Fecha_Liquidacion'] = df_deudas_fmt['Fecha_Liquidacion'].dt.strftime('%d/%m/%Y')

    st.success("¡Liquidación judicial finalizada!")
    avisar_tramos_sin_dias(avisos)
    st.markdown("### 📋 Resumen del Juicio (a los Intereses)")

    col1, col2, col3 = st.columns(3)
    with col1: st.metric("Intereses (Base del Juicio)", formato_arg(df_deudas['Capital'].sum()))
    with col2: st.metric("Resarcitorios", formato_arg(df_deudas['Interes_Resarcitorio'].sum()))
    with col3: st.metric("Punitorios", formato_arg(df_deudas['Interes_Punitorio'].sum()))

    col4, col5 = st.columns(2)
    with col4: st.metric("Total Actualizado", formato_arg(df_deudas['Total_Actualizado'].sum()))
    with col5: st.metric("Antigüedad Juicio", f"{antiguedad_juicio} días")

    st.divider()

    columnas_detalle = ['Impuesto', 'concepto', 'Periodo', 'Vencimiento', 'Capital',
                        'fecha_Demanda', 'Dias_Resarcitorios', 'Interes_Resarcitorio',
                        'Dias_Punitorios', 'Interes_Punitorio',
                        'Fecha_Liquidacion', 'Total_Actualizado']
    columnas_detalle = [c for c in columnas_detalle if c in df_deudas_fmt.columns]

    with st.expander("🔍 Ver detalle completo de obligaciones", expanded=False):
        st.dataframe(df_deudas_fmt[columnas_detalle], use_container_width=True)

    st.divider()
    mostrar_tabla_tasas(
        df_tasas_res, df_tasas_pun, meta,
        ventana_res=(df_deudas['Vencimiento'].min(), ultima_fecha_demanda),
        ventana_pun=(df_deudas['fecha_Demanda'].min(), ultima_fecha_liq),
        titulo_res="**Resarcitorios**")

    boton_descarga(
        df_deudas_fmt[columnas_detalle], "Liquidacion_ARCA_Intereses.xlsx", clave=clave,
        columnas_moneda=['Capital', 'Interes_Resarcitorio', 'Interes_Punitorio', 'Total_Actualizado'],
        columnas_totalizar=['Capital', 'Interes_Resarcitorio', 'Interes_Punitorio', 'Total_Actualizado'])

    with st.expander("🧾 Generar los VEPs para pagar esto"):
        generar_veps(df_deudas, clave)


# =====================================================================================
# GENERADOR DE PLANTILLAS EN BLANCO (para descargar y completar)
# =====================================================================================
_AZUL_HEADER = "1E3A8A"
_GRIS_EJEMPLO = "F0F4F8"
_FUENTE = "Arial"

def _estilo_header(ws, fila, columnas):
    for idx, col in enumerate(columnas, start=1):
        celda = ws.cell(row=fila, column=idx, value=col)
        celda.font = Font(name=_FUENTE, bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill(start_color=_AZUL_HEADER, end_color=_AZUL_HEADER, fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[fila].height = 30

def _fila_ejemplo(ws, fila, valores, formatos=None):
    for idx, val in enumerate(valores, start=1):
        celda = ws.cell(row=fila, column=idx, value=val)
        celda.font = Font(name=_FUENTE, italic=True, color="6B7280", size=10)
        celda.fill = PatternFill(start_color=_GRIS_EJEMPLO, end_color=_GRIS_EJEMPLO, fill_type="solid")
        if formatos and formatos[idx - 1]:
            celda.number_format = formatos[idx - 1]

def _ajustar_anchos(ws, anchos):
    for idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho

def _nota(ws, fila, rango, texto):
    ws.merge_cells(rango)
    ws.cell(row=fila, column=1, value=texto).font = Font(name=_FUENTE, italic=True, size=9, color="9CA3AF")

def _hoja_instrucciones(wb, titulo, lineas):
    ws = wb.create_sheet("Instrucciones", 0)
    ws["A1"] = titulo
    ws["A1"].font = Font(name=_FUENTE, bold=True, size=14, color=_AZUL_HEADER)
    for i, linea in enumerate(lineas, start=3):
        ws.cell(row=i, column=1, value=linea).font = Font(name=_FUENTE, size=11)
        ws.row_dimensions[i].height = 18
    ws.column_dimensions["A"].width = 110
    ws.sheet_view.showGridLines = False


def generar_plantilla_capital():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _hoja_instrucciones(wb, "Plantilla - Juicio por Capital + Intereses", [
        "Completá la hoja 'Deudas' con una fila por cada obligación (impuesto/período) del juicio.",
        "",
        "Columnas de 'Deudas':",
        "  • Impuesto: nombre del impuesto (ej: IMPUESTO A LAS GANANCIAS, IVA, etc.)",
        "  • concepto: detalle del concepto (ej: Saldo DDJJ, Anticipo, etc.)",
        "  • Periodo: período fiscal (ej: 2024-1)",
        "  • Vencimiento: fecha de vencimiento original de la obligación",
        "  • Capital: monto de capital adeudado (impuesto original, sin intereses)",
        "  • F. Pago Capital: fecha en que se pagó el capital. DEJAR VACÍO si el capital todavía no fue pagado.",
        "  • fecha_Demanda: fecha de inicio de la demanda de ejecución fiscal",
        "  • Fecha_Liquidacion: fecha a la que querés calcular la liquidación (fecha de pago de intereses)",
        "",
        "No hace falta cargar tasas: la app ya tiene la tabla oficial completa de ARCA, desde 1901",
        "hasta el tramo vigente, y se actualiza sola cuando ARCA publica una tasa nueva. Si tenés una",
        "planilla vieja con hojas 'Tasas', podés borrarlas: la app las ignora.",
        "",
        "Los días se cuentan como los cuenta ARCA: cada mes completo vale 30 días y el resto por días",
        "corridos. La app los calcula sola y los muestra en el detalle para que puedas auditarlos.",
        "",
        "Borrá la fila de ejemplo de 'Deudas' (en gris/cursiva) antes de cargar tus propios datos.",
    ])

    ws = wb.create_sheet("Deudas")
    _estilo_header(ws, 1, ["Impuesto", "concepto", "Periodo", "Vencimiento", "Capital", "F. Pago Capital", "fecha_Demanda", "Fecha_Liquidacion"])
    _fila_ejemplo(ws, 2,
                  ["IMPUESTO A LAS GANANCIAS", "Saldo DDJJ", "2024-1",
                   datetime.date(2024, 5, 20), 150000.00, None,
                   datetime.date(2025, 1, 15), datetime.date(2025, 6, 30)],
                  formatos=[None, None, None, "DD/MM/YYYY", "#,##0.00", "DD/MM/YYYY", "DD/MM/YYYY", "DD/MM/YYYY"])
    _ajustar_anchos(ws, [26, 20, 12, 14, 16, 16, 14, 16])
    ws["F1"].comment = Comment("Dejar vacío si el capital todavía no fue pagado.", "Liquidador ARCA")
    _nota(ws, 3, "A3:H3", "↑ Fila de ejemplo: borrala y cargá tus propias obligaciones (podés agregar tantas filas como necesites).")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generar_plantilla_intereses():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _hoja_instrucciones(wb, "Plantilla - Juicio a los Intereses", [
        "Usá esta planilla cuando el capital impositivo original YA fue pagado, y el juicio se inicia",
        "por los intereses resarcitorios que quedaron impagos (esos intereses pasan a ser la nueva",
        "'base' de la deuda). No incluye columna de pago de capital porque no aplica en este caso.",
        "",
        "Columnas de 'Deudas':",
        "  • Impuesto: nombre del impuesto (ej: IMPUESTO A LAS GANANCIAS, IVA, etc.)",
        "  • concepto: detalle del concepto (ej: INTERESES Capitalizables)",
        "  • Periodo: período fiscal de origen (ej: 2025-1)",
        "  • Vencimiento: fecha desde la que corren los intereses sobre este monto (vencimiento de origen)",
        "  • Capital: monto de intereses adeudados que se convierte en la base de este juicio",
        "  • fecha_Demanda: fecha de inicio de la demanda de ejecución fiscal",
        "  • Fecha_Liquidacion: fecha a la que querés calcular la liquidación (fecha de pago de intereses)",
        "",
        "No hace falta cargar tasas: la app ya tiene la tabla oficial completa de ARCA, desde 1901",
        "hasta el tramo vigente, y se actualiza sola cuando ARCA publica una tasa nueva. Si tenés una",
        "planilla vieja con hojas 'Tasas', podés borrarlas: la app las ignora.",
        "",
        "Los días se cuentan como los cuenta ARCA: cada mes completo vale 30 días y el resto por días",
        "corridos. La app los calcula sola y los muestra en el detalle para que puedas auditarlos.",
        "",
        "Borrá la fila de ejemplo de 'Deudas' (en gris/cursiva) antes de cargar tus propios datos.",
    ])

    ws = wb.create_sheet("Deudas")
    _estilo_header(ws, 1, ["Impuesto", "concepto", "Periodo", "Vencimiento", "Capital", "fecha_Demanda", "Fecha_Liquidacion"])
    _fila_ejemplo(ws, 2,
                  ["IMPUESTO A LAS GANANCIAS", "INTERESES Capitalizables", "2025-1",
                   datetime.date(2025, 6, 17), 3837980.63,
                   datetime.date(2026, 3, 13), datetime.date(2026, 7, 30)],
                  formatos=[None, None, None, "DD/MM/YYYY", "#,##0.00", "DD/MM/YYYY", "DD/MM/YYYY"])
    _ajustar_anchos(ws, [26, 24, 12, 14, 16, 14, 16])
    _nota(ws, 3, "A3:G3", "↑ Fila de ejemplo: borrala y cargá tus propias obligaciones (podés agregar tantas filas como necesites).")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# =====================================================================================
# IMPORTACIÓN DESDE EL MAIL DEL AGENTE FISCAL
# =====================================================================================
# El mail no se liquida directo: se convierte en planilla, y la planilla se sube a la
# pestaña que corresponda. Suena a paso de más y es a propósito. Así la importación
# usa el mismo camino de cálculo que ya está probado contra ARCA, y queda una planilla
# en el medio que se puede abrir, revisar y guardar en el expediente.

def _a_fecha(texto):
    """De 'AAAA-MM-DD' a date. None si viene vacío."""
    return datetime.date.fromisoformat(texto) if texto else None


def agentes_fiscales():
    """Las direcciones de los agentes fiscales que mandan las boletas.

    NO van en el código: el repositorio es público y son direcciones personales de
    gente con nombre y apellido. Se cargan en Streamlit Community Cloud, en
    Settings → Secrets, con esta forma:

        agentes_fiscales = "una@ejemplo.com, otra@ejemplo.com"

    Si no están cargadas, la app funciona igual pero no controla el remitente.
    """
    try:
        crudo = st.secrets.get("agentes_fiscales", "")
    except Exception:
        # Sin archivo de secrets configurado, st.secrets protesta. No es un error:
        # simplemente no hay lista.
        crudo = ""
    return [d.strip().lower() for d in str(crudo).split(",") if d.strip()]


def procesar_mail(archivo_subido):
    habilitados = agentes_fiscales()
    boleta = leer_mail.leer_boleta(archivo_subido.getvalue(), habilitados)

    # --- Lo que dice la carátula de la boleta ---
    st.markdown("#### Boleta de deuda")
    c1, c2, c3 = st.columns(3)
    c1.metric("Contribuyente", boleta['contribuyente'] or "—")
    c2.metric("CUIT", boleta['cuit'] or "—")
    c3.metric("Juicio", boleta['juicio'] or "—")

    for aviso in boleta['avisos']:
        st.warning(aviso)

    if not habilitados:
        st.caption(
            "⚠️ No hay cargada una lista de agentes fiscales, así que no controlé de quién "
            "vino el mail. Se carga en Settings → Secrets, con la clave `agentes_fiscales`."
        )
    elif boleta['remitente'] in habilitados:
        st.caption(f"Remitente reconocido: {boleta['remitente']}")

    df = pd.DataFrame(boleta['filas'])

    # --- Control automático: la suma de las filas contra el monto de demanda ---
    # Si estos dos números no coinciden, alguna fila se leyó mal o quedó afuera.
    # Es el mejor control que da el mail, porque el propio mail trae los dos.
    suma = round(df['Capital'].sum(), 2)
    declarado = boleta['monto_demanda']
    if declarado is None:
        st.info(f"Leí **{len(df)} filas** por **{formato_arg(suma)}**. "
                "La boleta no trae monto de demanda, así que no pude contrastarlo.")
    elif abs(suma - declarado) < 0.01:
        st.success(f"Leí **{len(df)} filas** por **{formato_arg(suma)}**, "
                   "que es exactamente el monto de demanda de la boleta.")
    else:
        st.error(
            f"**Las filas no suman el monto de demanda.** Leí {len(df)} filas por "
            f"{formato_arg(suma)} y la boleta declara {formato_arg(declarado)} "
            f"(diferencia: {formato_arg(round(declarado - suma, 2))}). "
            "Algo se leyó mal o quedó afuera: revisá el mail antes de usar esto."
        )

    # --- Las dos fechas que no salen del mail ---
    st.markdown("#### Fechas de la liquidación")
    c1, c2 = st.columns(2)
    fecha_demanda = c1.date_input(
        "Fecha de demanda", value=_a_fecha(boleta['fecha_demanda']) or datetime.date.today(),
        format="DD/MM/YYYY", key="mail_demanda",
        help="Viene de la fecha de sorteo de la boleta. Cambiala si corresponde otra.")
    fecha_liquidacion = c2.date_input(
        "Fecha de liquidación", value=datetime.date.today(), format="DD/MM/YYYY",
        key="mail_liquidacion", help="Hasta qué día se calculan los intereses. La elegís vos.")

    # --- Revisión fila por fila ---
    st.markdown("#### Revisá lo que leí")
    st.caption(
        "La columna **Destino** es una propuesta, armada según el subconcepto de ARCA: "
        "los saldos de declaración jurada y los anticipos son capital impositivo, y los "
        "intereses resarcitorios ya son interés. Corregila donde haga falta: las filas "
        "en «revisar» no entran en ninguna planilla hasta que las clasifiques."
    )

    df['Vencimiento'] = pd.to_datetime(df['Vencimiento'], errors='coerce')
    df['F. Pago Capital'] = pd.to_datetime(df['F. Pago Capital'], errors='coerce')

    editado = st.data_editor(
        df[['Destino', 'Impuesto', 'concepto', 'Periodo', 'Vencimiento', 'Capital',
            'F. Pago Capital', 'Nota', 'Aviso']],
        use_container_width=True, hide_index=True, num_rows="dynamic", key="mail_editor",
        column_config={
            'Destino': st.column_config.SelectboxColumn(
                "Destino", width="small", required=True,
                options=[leer_mail.CAPITAL, leer_mail.INTERESES, leer_mail.REVISAR],
                help="capital = se deben capital e intereses · intereses = el capital "
                     "está cancelado y se deben los intereses · revisar = no va a ninguna planilla"),
            'Vencimiento': st.column_config.DateColumn("Vencimiento", format="DD/MM/YYYY"),
            'F. Pago Capital': st.column_config.DateColumn("F. Pago Capital", format="DD/MM/YYYY"),
            'Capital': st.column_config.NumberColumn("Capital", format="%.2f"),
            'Nota': st.column_config.TextColumn("Nota del agente fiscal", width="medium", disabled=True),
            'Aviso': st.column_config.TextColumn("Por qué la marqué", width="medium", disabled=True),
        })

    pendientes = int((editado['Destino'] == leer_mail.REVISAR).sum())
    if pendientes:
        st.warning(f"{'Queda' if pendientes == 1 else 'Quedan'} **{pendientes} "
                   f"{'fila' if pendientes == 1 else 'filas'}** sin clasificar. "
                   "No van a salir en las planillas.")

    # --- Liquidar ---
    st.markdown("#### Liquidá")
    listo = editado.copy()
    listo['fecha_Demanda'] = pd.Timestamp(fecha_demanda)
    listo['Fecha_Liquidacion'] = pd.Timestamp(fecha_liquidacion)

    base = (boleta['contribuyente'] or 'boleta').title().replace(' ', '_')[:40]

    # La planilla se puede bajar —sirve para el expediente y para volver otro día—
    # pero no hace falta bajarla y volver a subirla: los datos ya están revisados acá.
    partes = []
    for destino, columnas, etiqueta, archivo in (
        (leer_mail.CAPITAL, COLUMNAS_CAPITAL, "Capital + Intereses", "Capital"),
        (leer_mail.INTERESES, COLUMNAS_INTERESES, "Juicio a los Intereses", "Intereses"),
    ):
        parte = listo[listo['Destino'] == destino]
        if not parte.empty:
            partes.append((parte, columnas, etiqueta, archivo))

    if not partes:
        st.info("No hay filas clasificadas todavía.")
        return

    columnas_ui = st.columns(len(partes))
    for col, (parte, columnas, etiqueta, archivo) in zip(columnas_ui, partes):
        resumen = (f"{len(parte)} {'fila' if len(parte) == 1 else 'filas'} · "
                   f"{formato_arg(round(parte['Capital'].sum(), 2))}")
        with col:
            st.markdown(f"**{etiqueta}** — {resumen}")
            st.checkbox("Liquidar acá mismo", key=f"liquidar_{archivo}")
            st.download_button(
                "⬇️ Bajar la planilla",
                data=armar_planilla(parte, columnas),
                file_name=f"{base}_{archivo}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"bajar_{archivo}")

    # Fuera de las columnas: la liquidación necesita el ancho completo, y además
    # trae su propio desplegable de VEPs, que no puede ir adentro de otro.
    for parte, columnas, etiqueta, archivo in partes:
        if not st.session_state.get(f"liquidar_{archivo}"):
            continue
        st.divider()
        st.markdown(f"### {etiqueta}")
        procesador = (procesar_juicio_capital if archivo == "Capital"
                      else procesar_juicio_intereses)
        try:
            procesador(parte[columnas], clave=f"mail_{archivo}")
        except Exception as e:
            st.error(f"No pude liquidar {etiqueta}: {e}")


# =====================================================================================
# GENERACIÓN DE VEPs
# =====================================================================================
# Una liquidación de ocho vencimientos con capital y los tres intereses son treinta y
# dos VEPs cargados de a uno. Con un archivo se cargan todos juntos.
#
# Nada se tilda solo: casi nunca se paga todo lo liquidado.

def _fecha_de(valor):
    """La liquidación baja las fechas como texto 'DD/MM/AAAA'."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    if isinstance(valor, datetime.date) and not isinstance(valor, datetime.datetime):
        return valor
    fecha = pd.to_datetime(valor, dayfirst=True, errors='coerce')
    return None if pd.isna(fecha) else fecha.date()


def procesar_veps(archivo_subido):
    """Entrada por archivo, para liquidaciones viejas que ya están en el disco."""
    df = pd.read_excel(archivo_subido)
    if 'Vencimiento' not in df.columns:
        raise ValueError(
            "El archivo no tiene la columna 'Vencimiento'. ¿Es una liquidación bajada "
            "de las otras pestañas?")
    generar_veps(df, "subido")


def generar_veps(df, clave):
    """La grilla de selección y el archivo.

    `clave` distingue las tres instancias que puede haber en pantalla a la vez
    (las dos liquidaciones y la pestaña de archivo), porque Streamlit necesita
    que cada control tenga un nombre propio.
    """
    # La liquidación trae una fila de totales al pie, que no es una obligación.
    df = df[df['Vencimiento'].notna() & df['Impuesto'].notna()].copy()
    if df.empty:
        raise ValueError("No encontré ninguna obligación en el archivo.")

    st.markdown("#### De quién es la deuda y quién paga")
    c1, c2, c3 = st.columns(3)
    cuit_contribuyente = c1.text_input(
        "CUIT del contribuyente", key=f"vep_cuit_contrib_{clave}", placeholder="30999999995",
        help="De quién es la deuda. Define si Ganancias va como sociedad o como persona física.")
    cuit_generador = c2.text_input(
        "CUIT del generador", key=f"vep_cuit_gen_{clave}", placeholder="20999999997",
        help="Quién sube el archivo a ARCA. Puede ser el mismo que el contribuyente.")
    fecha_exp = c3.date_input(
        "Vence el", value=datetime.date.today() + datetime.timedelta(days=veps.DIAS_EXPIRACION),
        format="DD/MM/YYYY", key=f"vep_expira_{clave}",
        help="Hasta cuándo se puede pagar el VEP. Por defecto, diez días.")

    if not cuit_contribuyente.strip():
        st.info("Cargá el CUIT del contribuyente para seguir.")
        return

    persona = veps.tipo_de_persona(cuit_contribuyente)
    if persona:
        st.caption(f"CUIT de persona **{persona}**"
                   + (" — Ganancias va como sociedad (10)." if persona == 'juridica'
                      else " — Ganancias va como persona física (11)."))
    else:
        st.warning(
            "Ese CUIT no empieza en un prefijo conocido (20, 23, 24 y 27 son personas "
            "físicas; 30, 33 y 34, jurídicas). Si hay Ganancias, no voy a poder decidir "
            "el código.")

    filas = df.to_dict('records')
    for fila in filas:
        fila['Vencimiento'] = _fecha_de(fila.get('Vencimiento'))
    candidatos = veps.preparar(filas, cuit_contribuyente)

    if not candidatos:
        st.warning("No hay ningún importe mayor a cero para pagar.")
        return

    # --- La grilla: un renglón por importe, todo destildado ---
    st.markdown("#### Elegí qué vas a pagar")
    st.caption(
        f"Cada importe de la liquidación es un VEP aparte. Hay **{len(candidatos)}** "
        "para elegir, y arrancan todos sin tildar.")

    tabla = pd.DataFrame([{
        'Pagar': False,
        'Concepto': {'Capital': 'Capital', 'Interes_Resarcitorio': 'Resarcitorios',
                     'Interes_Capitalizable': 'Capitalizables',
                     'Interes_Punitorio': 'Punitorios'}[c['columna']],
        'Impuesto': c['Impuesto'],
        'Vencimiento': c['Vencimiento'],
        'Período': c['periodo'],
        'Cuota': c['cuota'] or '',
        'Importe': c['importe'],
        'Form.': c['formulario'] or '',
        'Problema': c['aviso'],
    } for c in candidatos])

    editado = st.data_editor(
        tabla, use_container_width=True, hide_index=True, key=f"vep_editor_{clave}",
        column_config={
            'Pagar': st.column_config.CheckboxColumn("Pagar", width="small"),
            'Vencimiento': st.column_config.DateColumn("Vencimiento", format="DD/MM/YYYY", disabled=True),
            'Importe': st.column_config.NumberColumn("Importe", format="%.2f", disabled=True),
            'Problema': st.column_config.TextColumn("Problema", width="medium", disabled=True),
        },
        disabled=['Concepto', 'Impuesto', 'Período', 'Cuota', 'Form.'])

    elegidos = [c for c, tildado in zip(candidatos, editado['Pagar']) if tildado]

    # --- Los que no se pueden armar ---
    con_problema = [c for c in elegidos if not c['formulario']]
    if con_problema:
        st.error(
            f"**{len(con_problema)} de los tildados no se pueden armar** y no van a salir "
            "en el archivo. Mirá la columna «Problema»: hasta que eso se resuelva, esos "
            "importes se cargan a mano en ARCA.")

    listos = [c for c in elegidos if c['formulario']]
    total = round(sum(c['importe'] for c in listos), 2)

    c1, c2 = st.columns([1, 2])
    c1.metric("VEPs a generar", len(listos))
    c2.metric("Total a pagar", formato_arg(total))

    if not listos:
        st.info("Tildá al menos un importe que se pueda armar.")
        return

    if not cuit_generador.strip():
        st.info("Falta el CUIT del generador, que es quien sube el archivo a ARCA.")
        return

    try:
        contenido = veps.armar_txt(listos, cuit_generador, fecha_exp)
    except ValueError as e:
        st.error(str(e))
        return

    st.download_button(
        f"⬇️ Bajar el archivo ({len(listos)} VEPs · {formato_arg(total)})",
        data=contenido.encode('utf-8'),
        file_name=veps.nombre_archivo(cuit_generador),
        mime="text/plain", key=f"bajar_veps_{clave}")

    st.caption(
        "Se sube en ARCA, en **Presentación de DDJJ y Pagos → VEP → Generación masiva**. "
        "La primera vez, probá con un solo importe tildado: si ese entra, el resto también.")

    with st.expander("Ver el archivo antes de bajarlo"):
        st.code(contenido, language=None)


# =====================================================================================
# INTERFAZ
# =====================================================================================
# El acceso NO se controla acá: la app está marcada como privada en Streamlit
# Community Cloud, que pide login con cuenta de Google antes de servir nada. Quien no
# esté invitado no llega ni a esta línea.
#
# ⚠️ Si alguna vez se pasa la app a pública, o se muda a un hosting sin login propio,
# queda abierta a cualquiera con la URL. Hubo una puerta propia (acceso.py, con usuario
# y contraseña) que se sacó para no encadenar dos logins: está en el historial de git
# por si hace falta recuperarla.
tab0, tab1, tab2, tab3 = st.tabs(["📧 Importar desde el mail", "💰 Juicio por Capital + Intereses",
                                  "📈 Juicio a los Intereses", "🧾 Generar VEPs"])

with tab0:
    st.markdown(
        "Subí el mail del agente fiscal con la **boleta de deuda** y armo las planillas. "
        "Guardá el mail como archivo `.eml` (en Gmail: **⋮ → Descargar mensaje**) y arrastralo acá."
    )
    archivo_0 = st.file_uploader("Arrastrá el mail acá", type=["eml"], key="uploader_mail")
    if archivo_0 is not None:
        try:
            procesar_mail(archivo_0)
        except ValueError as e:
            st.error(str(e))
        except Exception as e:
            st.error(f"No pude leer el mail: {e}")

with tab1:
    st.markdown("Subí el Excel con la hoja **Deudas** (formato con Capital impositivo). Las tasas ya están en la app.")
    # El key del contenedor genera la clase .st-key-<key>, que es lo que engancha el CSS
    # para pintar de dorado el botón secundario (un div suelto no envolvería al widget).
    with st.container(key="lq-plantilla-capital"):
        st.download_button(
            label="📄 Descargar plantilla en blanco",
            data=generar_plantilla_capital(),
            file_name="Plantilla_Juicio_Capital_Intereses.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="plantilla_capital"
        )
    archivo_1 = st.file_uploader("Arrastrá tu Excel aquí", type=["xlsx"], key="uploader_capital")
    if archivo_1 is not None:
        try:
            with st.spinner('Procesando liquidación judicial...'):
                procesar_juicio_capital(archivo_1)
        except Exception as e:
            st.error(f"Error al procesar: {e}")

with tab2:
    st.markdown("Subí el Excel con la hoja **Deudas** (formato donde 'Capital' es el monto de intereses adeudados; el capital impositivo original ya está pago). Las tasas ya están en la app.")
    with st.container(key="lq-plantilla-intereses"):
        st.download_button(
            label="📄 Descargar plantilla en blanco",
            data=generar_plantilla_intereses(),
            file_name="Plantilla_Juicio_Intereses.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="plantilla_intereses"
        )
    archivo_2 = st.file_uploader("Arrastrá tu Excel aquí", type=["xlsx"], key="uploader_intereses")
    if archivo_2 is not None:
        try:
            with st.spinner('Procesando liquidación judicial...'):
                procesar_juicio_intereses(archivo_2)
        except Exception as e:
            st.error(f"Error al procesar: {e}")

with tab3:
    st.markdown(
        "Para una liquidación **de otro día**, que ya tenés guardada. Si acabás de "
        "calcularla, no hace falta pasar por acá: al pie del resultado, en las pestañas "
        "de al lado, está el mismo generador."
    )
    archivo_3 = st.file_uploader("Arrastrá la liquidación aquí", type=["xlsx"], key="uploader_veps")
    if archivo_3 is not None:
        try:
            procesar_veps(archivo_3)
        except ValueError as e:
            st.error(str(e))
        except Exception as e:
            st.error(f"No pude armar los VEPs: {e}")
