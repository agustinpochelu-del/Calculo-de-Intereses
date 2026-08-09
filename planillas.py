"""Armado de las planillas de deudas.

Está aparte de `app.py` para que la importación automática (`importar_boleta.py`)
pueda generar las mismas planillas sin arrastrar Streamlit, que en una tarea
programada no tiene dónde dibujar nada.

La forma de la hoja `Deudas` está definida en un solo lugar: acá. Si cambia una
columna, cambia para la app y para la rutina a la vez.
"""
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

AZUL_HEADER = "1E3A8A"
GRIS_EJEMPLO = "F0F4F8"
FUENTE = "Arial"

COLUMNAS_CAPITAL = ['Impuesto', 'concepto', 'Periodo', 'Vencimiento', 'Capital',
                    'F. Pago Capital', 'fecha_Demanda', 'Fecha_Liquidacion']
COLUMNAS_INTERESES = ['Impuesto', 'concepto', 'Periodo', 'Vencimiento', 'Capital',
                      'fecha_Demanda', 'Fecha_Liquidacion']


def estilo_header(ws, fila, columnas):
    for idx, col in enumerate(columnas, start=1):
        celda = ws.cell(row=fila, column=idx, value=col)
        celda.font = Font(name=FUENTE, bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill(start_color=AZUL_HEADER, end_color=AZUL_HEADER, fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[fila].height = 30


def ajustar_anchos(ws, anchos):
    for idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho


def armar_planilla(df, columnas):
    """Arma un Excel con la hoja 'Deudas', igual al que produce la plantilla."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deudas"
    estilo_header(ws, 1, columnas)

    for i, (_, fila) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(columnas, start=1):
            valor = fila.get(col)
            # Las fechas que faltan llegan como NaT y openpyxl no sabe escribirlas:
            # la celda tiene que quedar vacía.
            if valor is None or (hasattr(valor, '__class__') and str(valor) in ('NaT', 'nan')):
                valor = None
            celda = ws.cell(row=i, column=j, value=valor)
            if col == 'Capital':
                celda.number_format = '#,##0.00'
            elif 'echa' in col or col == 'Vencimiento':
                celda.number_format = 'DD/MM/YYYY'

    ajustar_anchos(ws, [26, 24, 12, 14, 16, 14, 14, 16][:len(columnas)])
    ws.freeze_panes = "A2"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
